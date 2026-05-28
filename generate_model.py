import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import LineChart, Reference

def create_model():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # 1. Sheet 1: Calculator
    ws1 = wb.create_sheet(title="Calculator")
    ws1.views.sheetView[0].showGridLines = True

    # Styling Palettes (Slate Blue theme)
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    section_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    input_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    font_title = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    font_section = Font(name=font_family, size=11, bold=True, color="1F4E79")
    font_label = Font(name=font_family, size=11, bold=False)
    font_bold = Font(name=font_family, size=11, bold=True)
    font_italic = Font(name=font_family, size=9, italic=True, color="595959")

    border_thin = Side(style='thin', color='D9D9D9')
    border_double = Side(style='double', color='1F4E79')
    border_thick = Side(style='medium', color='1F4E79')

    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    bottom_double_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_double)
    top_thick_border = Border(left=border_thin, right=border_thin, top=border_thick, bottom=border_thin)

    # Set Column Widths
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 45

    # Title Block
    ws1.merge_cells('A1:C1')
    ws1['A1'] = "GlowNest Amazon Campaign Profitability Calculator"
    ws1['A1'].font = font_title
    ws1['A1'].fill = header_fill
    ws1['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 40

    # Section Headers
    ws1['A3'] = "INPUT SECTION (Editable)"
    ws1['A3'].font = font_section
    ws1.merge_cells('A3:C3')
    ws1['A3'].fill = section_fill
    ws1['A3'].alignment = Alignment(vertical="center")
    ws1.row_dimensions[3].height = 24

    inputs = [
        ("Selling Price (SP)", 800, "₹#,##0.00", "Customer retail listing price on Amazon India"),
        ("Cost of Goods Sold (COGS)", 280, "₹#,##0.00", "Manufacturing + shipping to Amazon warehouse"),
        ("Amazon Referral Fee (%)", 0.12, "0.0%", "Category commission fee charged by Amazon"),
        ("Monthly Ad Spend", 300000, "₹#,##0", "Total advertising budget allocated for the month"),
        ("Units Sold via Ads", 1500, "#,##0", "Total volume of sales attributed directly to ads"),
        ("Organic Units Sold", 2000, "#,##0", "Total volume of sales generated from search engine rankings"),
        ("Fixed Costs (monthly)", 150000, "₹#,##0", "Warehousing, team payroll, software subscriptions")
    ]

    for idx, (label, value, num_format, note) in enumerate(inputs, start=4):
        ws1.cell(row=idx, column=1, value=label).font = font_label
        ws1.cell(row=idx, column=1).border = cell_border
        
        val_cell = ws1.cell(row=idx, column=2, value=value)
        val_cell.font = font_bold
        val_cell.number_format = num_format
        val_cell.alignment = Alignment(horizontal="right")
        val_cell.fill = input_fill
        val_cell.border = cell_border

        note_cell = ws1.cell(row=idx, column=3, value=note)
        note_cell.font = font_italic
        note_cell.border = cell_border
        ws1.row_dimensions[idx].height = 20

    # Outputs
    out_start_row = 12
    ws1.cell(row=out_start_row, column=1, value="OUTPUT SECTION (Auto-Calculated)").font = font_section
    ws1.merge_cells(f'A{out_start_row}:C{out_start_row}')
    ws1.cell(row=out_start_row, column=1).fill = section_fill
    ws1.cell(row=out_start_row, column=1).alignment = Alignment(vertical="center")
    ws1.row_dimensions[out_start_row].height = 24

    outputs = [
        ("Gross Margin per Unit", "=B4-B5", "₹#,##0.00", "= SP - COGS"),
        ("Amazon Fee per Unit", "=B4*B6", "₹#,##0.00", "= SP * Fee %"),
        ("Ad Cost per Unit", "=B7/B8", "₹#,##0.00", "= Ad Spend / Ad Units"),
        ("Net Margin (Ad Units)", "=B13-B14-B15", "₹#,##0.00", "= Gross Margin - Amazon Fee - Ad Cost per Unit"),
        ("Net Margin (Organic Units)", "=B13-B14", "₹#,##0.00", "= Gross Margin - Amazon Fee"),
        ("Total Monthly Revenue", "=B4*(B8+B9)", "₹#,##0", "= SP * (Ad Units + Organic Units)"),
        ("Total Monthly Profit", "=(B16*B8)+(B17*B9)-B10", "₹#,##0", "= (Net Margin Ad * Ad Units) + (Net Margin Org * Org Units) - Fixed Costs"),
        ("ACOS", "=B7/(B4*B8)", "0.0%", "= Ad Spend / Ad Revenue"),
        ("TACOS", "=B7/B18", "0.0%", "= Ad Spend / Total Revenue"),
        ("Break-Even Ad Spend", "=B17*(B8+B9)-B10", "₹#,##0", "Total contribution margin minus fixed costs")
    ]

    for idx, (label, formula, num_format, note) in enumerate(outputs, start=out_start_row+1):
        ws1.cell(row=idx, column=1, value=label).font = font_label
        ws1.cell(row=idx, column=1).border = cell_border
        
        val_cell = ws1.cell(row=idx, column=2, value=formula)
        val_cell.font = font_bold
        val_cell.number_format = num_format
        val_cell.alignment = Alignment(horizontal="right")
        val_cell.border = cell_border

        # Highlight important output cells
        if label in ["Total Monthly Profit", "Break-Even Ad Spend"]:
            val_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Light green highlight
            val_cell.border = bottom_double_border if label == "Total Monthly Profit" else cell_border
            ws1.cell(row=idx, column=1).font = font_bold
        elif label == "TACOS":
            val_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Light yellow default

        note_cell = ws1.cell(row=idx, column=3, value=note)
        note_cell.font = font_italic
        note_cell.border = cell_border
        ws1.row_dimensions[idx].height = 20

    # Conditional Formatting for Profit (Red if < 0) and TACOS (Yellow if > 15%)
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    red_font = Font(name=font_family, size=11, bold=True, color='9C0006')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    yellow_font = Font(name=font_family, size=11, bold=True, color='9C6500')

    ws1.conditional_formatting.add('B19', CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font))
    ws1.conditional_formatting.add('B21', CellIsRule(operator='greaterThan', formula=['0.15'], stopIfTrue=True, fill=yellow_fill, font=yellow_font))

    # 2. Sheet 2: Scenario Analysis
    ws2 = wb.create_sheet(title="Scenario Explorer")
    ws2.views.sheetView[0].showGridLines = True
    ws2.column_dimensions['A'].width = 32
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 16
    ws2.column_dimensions['E'].width = 16
    ws2.column_dimensions['F'].width = 30

    ws2.merge_cells('A1:F1')
    ws2['A1'] = "Scenario Analysis (Answers to Model Exploration)"
    ws2['A1'].font = font_title
    ws2['A1'].fill = header_fill
    ws2['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 40

    scenario_headers = ["Metric", "Base Case", "Price Sensitivity", "Cost Pressure", "Organic Growth", "Category Shift"]
    for col_idx, text in enumerate(scenario_headers, start=1):
        cell = ws2.cell(row=3, column=col_idx, value=text)
        cell.font = font_bold
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = cell_border
    ws2.row_dimensions[3].height = 24

    scenario_rows = [
        ("Selling Price (SP)", 800, 950, 800, 800, 700),
        ("COGS", 280, 280, 336, 280, 280),
        ("Amazon Referral Fee (%)", 0.12, 0.12, 0.12, 0.12, 0.12),
        ("Monthly Ad Spend", 300000, 300000, 300000, 300000, 300000),
        ("Units Sold via Ads", 1500, 1500, 1500, 1500, 1500),
        ("Organic Units Sold", 2000, 2000, 2000, 3500, 2000),
        ("Fixed Costs (monthly)", 150000, 150000, 150000, 150000, 150000)
    ]

    for r_idx, values in enumerate(scenario_rows, start=4):
        ws2.cell(row=r_idx, column=1, value=values[0]).font = font_label
        ws2.cell(row=r_idx, column=1).border = cell_border
        for c_idx, val in enumerate(values[1:], start=2):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_label
            cell.border = cell_border
            cell.alignment = Alignment(horizontal="right")
            if r_idx in [4, 5, 10]:
                cell.number_format = "₹#,##0"
            elif r_idx == 6:
                cell.number_format = "0.0%"
            else:
                cell.number_format = "#,##0"
        ws2.row_dimensions[r_idx].height = 20

    ws2.cell(row=11, column=1, value="Calculated Outputs").font = font_section
    ws2.merge_cells('A11:F11')
    ws2.cell(row=11, column=1).fill = section_fill
    ws2.row_dimensions[11].height = 24

    output_formulas = [
        ("Gross Margin per Unit", "=B4-B5", "₹#,##0.00"),
        ("Amazon Fee per Unit", "=B4*B6", "₹#,##0.00"),
        ("Ad Cost per Unit", "=B7/B8", "₹#,##0.00"),
        ("Net Margin (Ad Units)", "=B12-B13-B14", "₹#,##0.00"),
        ("Net Margin (Organic Units)", "=B12-B13", "₹#,##0.00"),
        ("Total Monthly Revenue", "=B4*(B8+B9)", "₹#,##0"),
        ("Total Monthly Profit", "=(B15*B8)+(B16*B9)-B10", "₹#,##0"),
        ("ACOS", "=B7/(B4*B8)", "0.0%"),
        ("TACOS", "=B7/B17", "0.0%"),
        ("Break-Even Ad Spend", "=B16*(B8+B9)-B10", "₹#,##0")
    ]

    for idx, (label, base_formula, num_format) in enumerate(output_formulas, start=12):
        ws2.cell(row=idx, column=1, value=label).font = font_label
        ws2.cell(row=idx, column=1).border = cell_border
        
        # Apply formulas to all columns B to F
        for c_idx, col_letter in enumerate(["B", "C", "D", "E", "F"], start=2):
            # Translate base formulas to dynamic column formulas
            # Base rows are: SP=4, COGS=5, Fee=6, Spend=7, AdUnits=8, OrgUnits=9, Fixed=10
            # Outputs are: Gross=12, FeePerUnit=13, AdCost=14, NetAd=15, NetOrg=16, Rev=17, Profit=18, ACOS=19, TACOS=20, BE=21
            f_map = {
                "Gross Margin per Unit": f"={col_letter}4-{col_letter}5",
                "Amazon Fee per Unit": f"={col_letter}4*{col_letter}6",
                "Ad Cost per Unit": f"={col_letter}7/{col_letter}8",
                "Net Margin (Ad Units)": f"={col_letter}12-{col_letter}13-{col_letter}14",
                "Net Margin (Organic Units)": f"={col_letter}12-{col_letter}13",
                "Total Monthly Revenue": f"={col_letter}4*({col_letter}8+{col_letter}9)",
                "Total Monthly Profit": f"=({col_letter}15*{col_letter}8)+({col_letter}16*{col_letter}9)-{col_letter}10",
                "ACOS": f"={col_letter}7/({col_letter}4*{col_letter}8)",
                "TACOS": f"={col_letter}7/{col_letter}17",
                "Break-Even Ad Spend": f"={col_letter}16*({col_letter}8+{col_letter}9)-{col_letter}10"
            }
            
            formula = f_map[label]
            val_cell = ws2.cell(row=idx, column=c_idx, value=formula)
            val_cell.font = font_bold if label in ["Total Monthly Profit", "Break-Even Ad Spend", "TACOS"] else font_label
            val_cell.number_format = num_format
            val_cell.alignment = Alignment(horizontal="right")
            val_cell.border = cell_border

            # Visual Highlights
            if label == "Total Monthly Profit":
                val_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                val_cell.border = bottom_double_border
            elif label == "Break-Even Ad Spend":
                val_cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            elif label == "TACOS":
                val_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        ws2.row_dimensions[idx].height = 20

    # Add red conditional format for profit column cells in Row 18 (B18 to F18)
    for col_letter in ["B", "C", "D", "E", "F"]:
        cell_ref = f"{col_letter}18"
        ws2.conditional_formatting.add(cell_ref, CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font))
        tacos_ref = f"{col_letter}20"
        ws2.conditional_formatting.add(tacos_ref, CellIsRule(operator='greaterThan', formula=['0.15'], stopIfTrue=True, fill=yellow_fill, font=yellow_font))

    # 3. Sheet 3: Spend-Profit Chart Data
    ws3 = wb.create_sheet(title="Chart Data & Viz")
    ws3.views.sheetView[0].showGridLines = True
    ws3.column_dimensions['A'].width = 16
    ws3.column_dimensions['B'].width = 16
    ws3.column_dimensions['C'].width = 16

    ws3.merge_cells('A1:C1')
    ws3['A1'] = "Ad Spend vs. Net Profit Analysis"
    ws3['A1'].font = font_title
    ws3['A1'].fill = header_fill
    ws3['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 40

    chart_headers = ["Monthly Ad Spend", "Total Monthly Profit", "TACOS"]
    for col_idx, text in enumerate(chart_headers, start=1):
        cell = ws3.cell(row=3, column=col_idx, value=text)
        cell.font = font_bold
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = cell_border
    ws3.row_dimensions[3].height = 24

    # Build simulated ad spend curve (Assuming diminishing returns on ad units)
    # Ad Units = 400 * ln(Ad Spend / 10000 + 1) + 300 (roughly modeling diminishing returns: at 3L spend -> ~1500 units)
    # Let's keep it simple for Excel formulas:
    # Let's write the values:
    spend_data = [
        (0, 0, 0),
        (50000, 480, 0.05),
        (100000, 850, 0.07),
        (150000, 1100, 0.08),
        (200000, 1280, 0.09),
        (250000, 1400, 0.10),
        (300000, 1500, 0.107), # Base Case
        (400000, 1680, 0.12),
        (500000, 2100, 0.14),
        (600000, 2500, 0.167), # Scale Case
        (750000, 2750, 0.20),
        (900000, 2900, 0.23),
        (1200000, 3100, 0.30)
    ]

    for idx, (spend, ad_units, dummy_tacos) in enumerate(spend_data, start=4):
        ws3.cell(row=idx, column=1, value=spend).number_format = "₹#,##0"
        
        # We can reference the inputs of the Calculator sheet to calculate the profitability dynamically
        # Formula for Profit at this specific spend:
        # Net Margin Ad = (Gross Margin - Amazon Fee) - Spend/AdUnits
        # Net Margin Org = Gross Margin - Amazon Fee
        # Profit = Net Margin Ad * AdUnits + Net Margin Org * OrgUnits - Fixed
        # Which simplifes to: (Net Margin Org * Total Units) - Spend - Fixed
        # Total Units = AdUnits + OrgUnits
        # Profit = Calculator!B16 * (AdUnits + Calculator!B8) - Spend - Calculator!B9
        ws3.cell(row=idx, column=2, value=f"=Calculator!$B$17*({ad_units}+Calculator!$B$9)-A{idx}-Calculator!$B$10").number_format = "₹#,##0"
        
        # TACOS Formula = Spend / Total Revenue = Spend / (SP * Total Units)
        # =A{idx} / (Calculator!$B$4 * ({ad_units} + Calculator!$B$9))
        ws3.cell(row=idx, column=3, value=f"=A{idx}/(Calculator!$B$4*({ad_units}+Calculator!$B$9))").number_format = "0.0%"

        for c_idx in range(1, 4):
            ws3.cell(row=idx, column=c_idx).font = font_label
            ws3.cell(row=idx, column=c_idx).border = cell_border
            ws3.cell(row=idx, column=c_idx).alignment = Alignment(horizontal="right")
        ws3.row_dimensions[idx].height = 20

    # Add Line Chart
    chart = LineChart()
    chart.title = "Effect of Ad Spend on Total Monthly Profit"
    chart.style = 13
    chart.y_axis.title = "Net Profit (INR)"
    chart.x_axis.title = "Ad Spend (INR)"
    chart.width = 18
    chart.height = 10

    # Data is columns B (Profit), X-axis categories is column A (Spend)
    data = Reference(ws3, min_col=2, min_row=3, max_row=16)
    cats = Reference(ws3, min_col=1, min_row=4, max_row=16)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Style Chart Line
    s1 = chart.series[0]
    s1.graphicalProperties.line.solidFill = "1F4E79"
    s1.graphicalProperties.line.width = 30000  # thickness

    ws3.add_chart(chart, "E4")

    # Save Workbook
    out_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(out_dir, "GlowNest_Profitability_Model.xlsx")
    wb.save(file_path)
    print(f"Spreadsheet generated successfully at: {file_path}")

if __name__ == "__main__":
    create_model()
